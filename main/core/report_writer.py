import csv
import re
from io import BytesIO
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo


REPORT_COLUMNS = [
    "run_timestamp",
    "username",
    "first_name",
    "last_name",
    "dob",
    "debit_card",
    "cvv",
    "account_type",
    "initial_deposit",
    "loan_amount_usd",
    "down_payment_usd",
    "loan_amount_eur",
    "down_payment_eur",
    "usd_to_eur_rate",
    "rate_source",
    "report_currency",
    "register_status",
    "open_account_status",
    "loan_status",
    "logout_status",
    "overall_status",
    "failed_step",
    "error_message",
]


_ILLEGAL_XLSX_CHARS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


# Isvalo neleistinus simbolius is langelio reiksmes
def _sanitize_cell_value(value):
    if isinstance(value, str):
        return _ILLEGAL_XLSX_CHARS.sub("", value)
    return value


# Nustato i kurias vietas bus saugoma ataskaita
def _resolve_output_dirs(settings) -> list[Path]:
    report_dir = Path(settings.report_output_dir)
    if settings.desktop_report:
        desktop_dir = Path.home() / "Desktop"
        if desktop_dir.resolve() == report_dir.resolve():
            return [desktop_dir]
        return [desktop_dir, report_dir]
    return [report_dir]


# Iraso CSV ataskaita i visus nurodytus kelius
def _write_csv_report(rows: list[dict], output_paths: list[Path]) -> str:
    for output_path in output_paths:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with output_path.open("w", encoding="utf-8", newline="") as file:
            writer = csv.DictWriter(file, fieldnames=REPORT_COLUMNS, quoting=csv.QUOTE_MINIMAL)
            writer.writeheader()
            for row in rows:
                writer.writerow(row)
    return output_paths[0].as_posix()


# Pritaiko stulpeliu plocius pagal turini
def _fit_worksheet_columns(worksheet) -> None:
    for column in worksheet.columns:
        max_len = 0
        column_letter = column[0].column_letter
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        width = min(max(max_len + 2, 12), 60)
        worksheet.column_dimensions[column_letter].width = width


# Sukuria ir iraso suformatuota XLSX ataskaita
def _write_xlsx_report(rows: list[dict], output_paths: list[Path]) -> str:
    # Sukuria tuscia workbook ir uzdeda antrastes
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Report"
    worksheet.append(REPORT_COLUMNS)

    # Uzpildo eilutes ir isvalo neleistinus simbolius
    for row in rows:
        worksheet.append([_sanitize_cell_value(row.get(column, "")) for column in REPORT_COLUMNS])

    # Paryskina ir centruoja antrastes
    header_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Nustato paprasta lygiavima duomenu eilutems
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Skaiciu stulpeliams uzdeda vienoda 2 skaitmenu formata
    currency_fields = {"initial_deposit", "loan_amount_usd", "down_payment_usd", "loan_amount_eur", "down_payment_eur"}
    rate_fields = {"usd_to_eur_rate"}
    for cell in worksheet[1]:
        if cell.value in currency_fields or cell.value in rate_fields:
            column_index = cell.column
            for data_cell in worksheet.iter_cols(min_col=column_index, max_col=column_index, min_row=2, max_row=worksheet.max_row):
                for item in data_cell:
                    item.number_format = "0.00"

    # Uzfiksuoja antraste ir prideda lentele su stiliumi
    worksheet.freeze_panes = "A2"
    if worksheet.max_row >= 2:
        report_table = Table(displayName="ParabankReport", ref=worksheet.dimensions)
        report_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(report_table)
    else:
        worksheet.auto_filter.ref = worksheet.dimensions

    # Pritaiko stulpeliu plocius pagal ilgiausias reiksmes
    _fit_worksheet_columns(worksheet)

    # Issaugo viena karta i buferi ir iraso i visus kelius
    buffer = BytesIO()
    workbook.save(buffer)
    payload = buffer.getvalue()
    for output_path in output_paths:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(payload)
    return output_paths[0].as_posix()


# Parenka ataskaitos formata ir issaugo faila
def write_report(rows: list[dict], settings) -> str:
    report_type = str(settings.report_type).strip().lower()
    timestamp = datetime.now().strftime("%Y-%m-%d-%H%M%S")
    output_paths = [
        output_dir / f"parabank_report_{timestamp}.{report_type if report_type == 'xlsx' else 'csv'}"
        for output_dir in _resolve_output_dirs(settings)
    ]
    if report_type == "xlsx":
        return _write_xlsx_report(rows, output_paths)
    return _write_csv_report(rows, output_paths)
